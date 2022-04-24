# -*- coding: utf-8 -*-
"""
    USE: python bif_exel2brical.py infile outfile
"""
import sys
import math
import json
import openpyxl


def createModules(ws):
    modules = {}
    for p in range(2):
        for i in range(ws.max_row - 1):
            val = ws.cell(row=i + 2, column=1).value
            if val is not None and val.strip() != "":
                name = val.strip().replace(' ', '_').replace(':', '.')
                clm4 = ws.cell(row=i + 2, column=5).value  # Functionality
                if clm4 is None:
                    functionality = ""
                else:
                    functionality = clm4.strip()
                clm6 = ws.cell(row=i + 2, column=7).value  # implementation
                if clm6 is None:
                    implClass = ""
                else:
                    implClass = clm6.strip()
                labels = ws.cell(row=i + 2, column=2).value
                if labels is None:
                    labels = ""
                parts = ws.cell(row=i + 2, column=4).value
                submodules = []
                if parts is not None:
                    submodules = parts.split(';')
                    for j in range(len(submodules)):
                        submodules[j] = submodules[j].strip()
                module = {"Name": name, "Comment": labels + ":" + functionality}
                if len(submodules) > 0:
                    module["SubModules"] = submodules
                else:
                    module["ImplClass"] = implClass
                modules[name] = module
    return modules


def upper_p(module1, module2, modules):
    if 'SubModules' in modules[module1]:
        if module2 in modules[module1]['SubModules']:
            return True
        else:
            for submodule in modules[module1]['SubModules']:
                if upper_p(submodule, module2, modules):
                    return True
    return False


def createConnections(ws, modules):
    connections = []
    ports = []
    for i in range(ws.max_row - 1):
        fromCircuit = ""
        col1 = ws.cell(row=i + 2, column=1).value   # fromCircuit
        if col1 is not None:
            col1 = col1.strip()
            fromCircuit = col1.strip().replace(' ', '_').replace(':', '.')
        fromPort = ""
        col2 = ws.cell(row=i + 2, column=2).value   # fromPort
        if col2 is not None:
            fromPort = col2.strip()
        toCircuit = ""
        col3 = ws.cell(row=i + 2, column=3).value   # toCircuit
        if col3 is not None:
            col3 = col3.strip()
            toCircuit = col3.strip().replace(' ', '_').replace(':', '.')
        toPort = ""
        col4 = ws.cell(row=i + 2, column=4).value   # toPort
        if col4 is not None:
            toPort = col4.strip()
        if fromCircuit == "" or toCircuit == "":
            continue
        if fromCircuit not in modules:
            sys.stderr.write("WARNING: " + fromCircuit + "is not defined in the Circuit sheet!\n")
        if toCircuit not in modules:
            sys.stderr.write("WARNING: " + toCircuit + "is not defined in the Circuit sheet!\n")
        connectionID = fromCircuit + "-" + toCircuit
        shape = []
        col5 = ws.cell(row=i + 2, column=5).value   # shape
        if col5 is not None:
            col5 = str(col5).strip()
            shape = col5.split(",")
            for j in range(len(shape)):
                try:
                    shape[j] = math.floor(float(shape[j]))
                except ValueError:
                    sys.stderr.write("WARNING: the shape element in " + connectionID + "is not an integer!\n")
        connection = {"Name": connectionID, "FromModule": fromCircuit, "FromPort": fromPort,
                      "ToModule": toCircuit, "ToPort": toPort}
        connections.append(connection)
        ports = add_ports(connection, shape, modules, ports)
    return connections, ports


def add_ports(connection, shape, modules, ports):
    fromModule = connection["FromModule"]
    toModule = connection["ToModule"]
    if upper_p(fromModule, toModule, modules):
        fromType = "Input"
        toType = "Input"
    elif upper_p(toModule, fromModule, modules):
        fromType = "Output"
        toType = "Output"
    else:
        fromType = "Output"
        toType = "Input"
    fromPort = {"Name": connection["FromPort"], "Type": fromType, "Shape": shape}
    if "Ports" in modules[fromModule]:
        if not defined_port(fromPort, modules[fromModule]["Ports"]):
            modules[fromModule]["Ports"].append(fromPort)
    else:
        modules[fromModule]["Ports"] = [fromPort]
    toPort = {"Name": connection["ToPort"], "Type": toType, "Shape": shape}
    if "Ports" in modules[toModule]:
        if not defined_port(toPort, modules[toModule]["Ports"]):
            modules[toModule]["Ports"].append(toPort)
    else:
        modules[toModule]["Ports"] = [toPort]
    return ports


def defined_port(port_2B_checked, ports):
    for port in ports:
        if port_2B_checked["Name"] == port["Name"] and port_2B_checked["Type"] == port["Type"]:
            return True
    return False


def main():
    if len(sys.argv) <= 2:
        print("USE: python bif_exel2brical.py infile bifd_url outfile")
        exit()

    outfilePath = sys.argv[2]
    wb = openpyxl.load_workbook(sys.argv[1])

    # Defining an ontology
    project = wb['Project']
    pname = project.cell(row=2, column=1).value
    if not pname:
        print("Error: no project name")
        exit()
    description = project.cell(row=2, column=3).value

    modules = createModules(wb['Circuit'])
    connections, ports = createConnections(wb['BriCA'], modules)

    module_array = []
    for v in modules.values():
        v["Ports"] = sorted(v["Ports"], key=lambda x: (x['Type'], x['Name']))
        module_array.append(v)

    output = {"Header": {"Type": "A", "Name": pname, "Base": pname, "Comment": description},
              "Modules": module_array,
              "Connections": connections}

    fp = open(outfilePath, 'w')
    json.dump(output, fp, indent=1)
    fp.close()


if __name__ == '__main__':
    main()
